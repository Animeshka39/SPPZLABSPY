from random import randint
import random
import string
import datetime
import openpyxl


# function that returns a number of the specific type you want to have, and with random values
def generateElement(type):
  if type == 1:
    return randint(0, 100)

  if type == 2:
    return random.random()

  if type == 3:
    return random.choice(string.ascii_letters)

  if type == 4:
    generatedType = randint(1, 3)
    return generateElement(generatedType)


print("Введите колличество елементов в массиве")
numberOfElements = int(input())
print("Какой тип елементов хотите добавить?")
print("Целые числа: 1, Дробные числа: 2, Символы: 3, Смешаный тип: 4")
typeOfElements = int(input())
i = 0
elements = []


start = datetime.datetime.now()
while i < numberOfElements:
  elements.append(generateElement(typeOfElements))
  i += 1
  if i == numberOfElements:
    break
finish = datetime.datetime.now()
duration = finish - start
print("Время выполнения: ", duration)
# openpyxl is here
path = "sample.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

sheet['A' + str(sheet.max_row + 1)] = sheet.max_row
sheet['B' + str(sheet.max_row)] = "Python"
sheet['C' + str(sheet.max_row)] = str(numberOfElements)
sheet['D' + str(sheet.max_row)] = typeOfElements
sheet['E' + str(sheet.max_row)] = str(duration)

wb.save('sample.xlsx')